import logging
import os
import sqlite3
from datetime import datetime, timedelta, date
from io import BytesIO
from typing import Optional, Dict, Any, List

import json
import requests
import pandas as pd
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from dotenv import load_dotenv

from telegram import (
    Update,
    ReplyKeyboardMarkup,
    InlineKeyboardMarkup,
    InlineKeyboardButton,
    InputFile,
)
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    CallbackQueryHandler,
    ContextTypes,
    filters,
)

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

AnyType = Any

# ----------------- ЛОГИ -----------------
logging.basicConfig(
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    level=logging.INFO,
)
log = logging.getLogger("sot_bot")

# ----------------- НАСТРОЙКИ И .ENV -----------------
load_dotenv()

BOT_TOKEN = (os.getenv("BOT_TOKEN") or "").strip()
DB_PATH = os.getenv("DB_PATH", "sot_bot.db")

TIMEZONE_OFFSET = int(os.getenv("TIMEZONE_OFFSET", "3"))
ANALYTICS_PASSWORD = "051995"


def _extract_spreadsheet_id_from_url(url: str) -> str:
    try:
        if "/d/" in url:
            return url.split("/d/")[1].split("/")[0]
    except Exception:
        pass
    return ""


SCHEDULE_URL_ENV = (os.getenv("SCHEDULE_URL") or "").strip()

_default_sheet_id = _extract_spreadsheet_id_from_url(SCHEDULE_URL_ENV)
if not _default_sheet_id:
    _default_sheet_id = (os.getenv("GSHEETS_SPREADSHEET_ID") or "").strip()
if not _default_sheet_id:
    _default_sheet_id = "1W_9Cs-LaX6KR4cE9xN71CliE6Lm_TyQqk8t3kQa4FCc"

GSHEETS_SPREADSHEET_ID = _default_sheet_id

if SCHEDULE_URL_ENV:
    GOOGLE_SHEET_URL_DEFAULT = SCHEDULE_URL_ENV
else:
    GOOGLE_SHEET_URL_DEFAULT = (
        f"https://docs.google.com/spreadsheets/d/{GSHEETS_SPREADSHEET_ID}/edit?usp=sharing"
    )

GSHEETS_SERVICE_ACCOUNT_JSON = (os.getenv("GSHEETS_SERVICE_ACCOUNT_JSON") or "").strip()
SHEETS_SERVICE = None

DEFAULT_APPROVERS = [
    "@asdinamitif",
    "@FrolovAlNGSN",
    "@cappit_G59",
    "@sergeybektiashkin",
    "@scri4",
    "@Kirill_Victorovi4",
]

RESPONSIBLE_USERNAMES: Dict[str, List[str]] = {
    "бектяшкин": ["sergeybektiashkin"],
    "смирнов": ["scri4"],
}

INSPECTOR_SHEET_NAME = "ПБ, АР,ММГН, АГО (2025)"
HARD_CODED_ADMINS = {398960707}

SCHEDULE_NOTIFY_CHAT_ID_ENV = (os.getenv("SCHEDULE_NOTIFY_CHAT_ID") or "").strip()
SCHEDULE_NOTIFY_CHAT_ID = (
    int(SCHEDULE_NOTIFY_CHAT_ID_ENV) if SCHEDULE_NOTIFY_CHAT_ID_ENV else None
)

# ВТОРАЯ ТАБЛИЦА — итоговые проверки
FINAL_CHECKS_SPREADSHEET_ID = (
    os.getenv(
        "FINAL_CHECKS_SPREADSHEET_ID",
        "1dUO3neTKzKI3D8P6fs_LJLmWlL7jw-FhohtJkjz4KuE",
    ).strip()
)


def is_admin(uid: int) -> bool:
    return uid in HARD_CODED_ADMINS


def local_now() -> datetime:
    return datetime.utcnow() + timedelta(hours=TIMEZONE_OFFSET)


def get_current_remarks_sheet_name() -> str:
    year = local_now().year
    return f"ПБ, АР,ММГН, АГО ({year})"


# -------------------------------------------------
# Google Sheets helpers
# -------------------------------------------------
def get_sheets_service():
    global SHEETS_SERVICE

    if SHEETS_SERVICE is not None:
        return SHEETS_SERVICE

    if not GSHEETS_SERVICE_ACCOUNT_JSON:
        log.error(
            "GSHEETS_SERVICE_ACCOUNT_JSON не задан – Google Sheets API недоступен."
        )
        return None

    try:
        info = json.loads(GSHEETS_SERVICE_ACCOUNT_JSON)
        creds = Credentials.from_service_account_info(
            info,
            scopes=["https://www.googleapis.com/auth/spreadsheets"],
        )
        service = build("sheets", "v4", credentials=creds)
        SHEETS_SERVICE = service
        return service
    except Exception as e:
        log.error("Ошибка создания клиента Google Sheets: %s", e)
        return None


def build_export_url(spreadsheet_id: str) -> str:
    return f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=xlsx"


def detect_header_row(values: List[List[str]]) -> int:
    for i, row in enumerate(values[:30]):
        row_lower = [str(c).lower() for c in row]
        if any("дата выезда" in c for c in row_lower):
            return i
    return 0


def read_sheet_to_dataframe(
    sheet_id: str, sheet_name: str, header_row_index: Optional[int] = None
) -> Optional[pd.DataFrame]:
    service = get_sheets_service()
    if service is None:
        log.error("Google Sheets сервис недоступен – невозможно прочитать лист.")
        return None

    try:
        result = (
            service.spreadsheets()
            .values()
            .get(spreadsheetId=sheet_id, range=f"'{sheet_name}'!A1:ZZZ1000")
            .execute()
        )
        values = result.get("values", [])

        if not values:
            log.warning("Лист '%s' пуст.", sheet_name)
            return pd.DataFrame()

        if header_row_index is None:
            header_row_index = detect_header_row(values)

        headers = values[header_row_index]
        data_rows = values[header_row_index + 1 :]

        df = pd.DataFrame(data_rows, columns=headers)
        df = df.dropna(how="all").reset_index(drop=True)
        return df
    except Exception as e:
        log.error("Ошибка чтения листа '%s' из Google Sheets: %s", sheet_name, e)
        return None


# -------------------------------------------------
# Работа со столбцами Excel
# -------------------------------------------------
def excel_col_to_index(col: str) -> int:
    col = col.upper().strip()
    idx = 0
    for ch in col:
        if "A" <= ch <= "Z":
            idx = idx * 26 + (ord(ch) - ord("A") + 1)
    return idx - 1


def get_col_by_letter(df: pd.DataFrame, letters: str) -> Optional[str]:
    idx = excel_col_to_index(letters)
    if 0 <= idx < len(df.columns):
        return df.columns[idx]
    return None


def get_col_index_by_header(
    df: pd.DataFrame, search_substr: str, fallback_letter: str
) -> Optional[int]:
    search_substr = search_substr.lower()
    for i, col in enumerate(df.columns):
        if search_substr in str(col).lower():
            return i
    idx = excel_col_to_index(fallback_letter)
    if 0 <= idx < len(df.columns):
        return idx
    return None


def normalize_onzs_value(val) -> Optional[str]:
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    try:
        n = int(float(s.replace(",", ".")))
        return str(n)
    except Exception:
        pass
    return s


def normalize_case_number(val) -> str:
    """
    Нормализация номера дела:

    - приводим все нестандартные тире к обычному '-';
    - убираем пробелы;
    - выбрасываем любые символы, кроме цифр и '-'.

    Примеры:
    'Дело № 03–46–108600 (ПП)' -> '03-46-108600'
    ' 01-29-099900 ' -> '01-29-099900'
    """
    if val is None:
        return ""
    s = str(val).strip()
    if not s:
        return ""

    # все «косые» тире в нормальное
    hyphens = ["\u2010", "\u2011", "\u2012", "\u2013", "\u2014", "\u2212"]
    for h in hyphens:
        s = s.replace(h, "-")

    # убираем пробелы
    s = s.replace(" ", "")

    # оставляем только цифры и '-'
    cleaned_chars = []
    for ch in s:
        if ch.isdigit() or ch == "-":
            cleaned_chars.append(ch)

    return "".join(cleaned_chars)


def get_case_col_index(df: pd.DataFrame) -> Optional[int]:
    idx_i = excel_col_to_index("I")
    if 0 <= idx_i < len(df.columns):
        return idx_i
    return get_col_index_by_header(df, "номер дела", "I")


# -------------------------------------------------
# БАЗА ДАННЫХ
# -------------------------------------------------
def get_db() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db() -> None:
    conn = get_db()
    c = conn.cursor()

    c.execute(
        """CREATE TABLE IF NOT EXISTS schedule_settings (
               key TEXT PRIMARY KEY,
               value TEXT
           )"""
    )

    c.execute(
        """CREATE TABLE IF NOT EXISTS approvers (
               id INTEGER PRIMARY KEY AUTOINCREMENT,
               label TEXT UNIQUE
           )"""
    )

    c.execute(
        """CREATE TABLE IF NOT EXISTS schedule_files (
               version INTEGER PRIMARY KEY,
               name TEXT,
               uploaded_at TEXT
           )"""
    )

    c.execute(
        """CREATE TABLE IF NOT EXISTS schedule_approvals (
               id INTEGER PRIMARY KEY AUTOINCREMENT,
               version INTEGER,
               approver TEXT,
               status TEXT,
               comment TEXT,
               decided_at TEXT,
               requested_at TEXT
           )"""
    )

    c.execute(
        """CREATE TABLE IF NOT EXISTS inspector_visits (
               id INTEGER PRIMARY KEY AUTOINCREMENT,
               date TEXT,
               area TEXT,
               floors TEXT,
               onzs TEXT,
               developer TEXT,
               object TEXT,
               address TEXT,
               case_no TEXT,
               check_type TEXT,
               created_at TEXT
           )"""
    )

    c.execute("SELECT COUNT(*) AS c FROM approvers")
    if c.fetchone()["c"] == 0:
        c.executemany(
            "INSERT OR IGNORE INTO approvers (label) VALUES (?)",
            [(lbl,) for lbl in DEFAULT_APPROVERS],
        )

    c.execute("SELECT value FROM schedule_settings WHERE key='schedule_version'")
    if not c.fetchone():
        c.execute(
            "INSERT INTO schedule_settings (key, value) VALUES ('schedule_version', '1')"
        )

    c.execute("SELECT value FROM schedule_settings WHERE key='last_notified_version'")
    if not c.fetchone():
        c.execute(
            "INSERT INTO schedule_settings (key, value) VALUES ('last_notified_version', '0')"
        )

    if SCHEDULE_NOTIFY_CHAT_ID_ENV:
        c.execute(
            "INSERT OR REPLACE INTO schedule_settings (key, value) VALUES (?, ?)",
            ("schedule_notify_chat_id", SCHEDULE_NOTIFY_CHAT_ID_ENV),
        )

    conn.commit()
    conn.close()


def get_schedule_state() -> dict:
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT key, value FROM schedule_settings")
    rows = c.fetchall()
    conn.close()
    return {r["key"]: r["value"] for r in rows}


def get_schedule_version(settings: dict) -> int:
    try:
        return int(settings.get("schedule_version") or "1")
    except Exception:
        return 1


def get_current_approvers(settings: dict) -> List[str]:
    val = settings.get("current_approvers")
    if val:
        arr = [v.strip() for v in val.split(",") if v.strip()]
        if arr:
            return arr
    return []


def set_current_approvers_for_version(approvers: List[str], version: int) -> None:
    conn = get_db()
    c = conn.cursor()

    c.execute(
        "INSERT OR REPLACE INTO schedule_settings (key, value) VALUES ('current_approvers', ?)",
        (",".join(approvers),),
    )

    c.execute("DELETE FROM schedule_approvals WHERE version = ?", (version,))

    now = local_now().isoformat()
    for appr in approvers:
        c.execute(
            """INSERT INTO schedule_approvals
               (version, approver, status, comment, decided_at, requested_at)
               VALUES (?, ?, 'pending', NULL, NULL, ?)""",
            (version, appr, now),
        )

    conn.commit()
    conn.close()


def get_schedule_approvals(version: int) -> List[sqlite3.Row]:
    conn = get_db()
    c = conn.cursor()
    c.execute(
        "SELECT * FROM schedule_approvals WHERE version = ? ORDER BY approver",
        (version,),
    )
    rows = c.fetchall()
    conn.close()
    return rows


def update_schedule_approval_status(
    version: int, approver: str, status: str, comment: Optional[str] = None
):
    conn = get_db()
    c = conn.cursor()
    now = local_now().isoformat()

    c.execute(
        """UPDATE schedule_approvals
           SET status=?, comment=?, decided_at=?
           WHERE version=? AND approver=?""",
        (status, comment, now, version, approver),
    )
    conn.commit()
    conn.close()


# -------------------------------------------------
# Инспектор: БД
# -------------------------------------------------
def save_inspector_to_db(form: Dict[str, Any]) -> bool:
    try:
        conn = get_db()
        c = conn.cursor()
        date_obj = form.get("date")
        date_str = date_obj.strftime("%Y-%m-%d") if date_obj else None
        c.execute(
            """INSERT INTO inspector_visits
               (date, area, floors, onzs, developer, object, address,
                case_no, check_type, created_at)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                date_str,
                form.get("area", ""),
                form.get("floors", ""),
                form.get("onzs", ""),
                form.get("developer", ""),
                form.get("object", ""),
                form.get("address", ""),
                form.get("case", ""),
                form.get("check_type", ""),
                local_now().isoformat(),
            ),
        )
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        log.error("Ошибка сохранения инспектора в локную БД: %s", e)
        return False


def fetch_inspector_visits(limit: int = 50) -> List[sqlite3.Row]:
    conn = get_db()
    c = conn.cursor()
    c.execute(
        """SELECT * FROM inspector_visits
           ORDER BY date DESC, id DESC
           LIMIT ?""",
        (limit,),
    )
    rows = c.fetchall()
    conn.close()
    return rows


def clear_inspector_visits() -> None:
    conn = get_db()
    c = conn.cursor()
    c.execute("DELETE FROM inspector_visits")
    conn.commit()
    conn.close()


# -------------------------------------------------
# Клавиатуры
# -------------------------------------------------
def main_menu() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        [
            ["📅 График", "📝 Замечания"],
            ["Инспектор", "📈 Аналитика"],
            ["Итоговые проверки"],
        ],
        resize_keyboard=True,
    )


def build_schedule_inline(
    is_admin_flag: bool, settings: dict, user_tag: Optional[str] = None
) -> InlineKeyboardMarkup:
    buttons = [
        [
            InlineKeyboardButton("🔄 Обновить", callback_data="schedule_refresh"),
            InlineKeyboardButton("📥 Скачать", callback_data="schedule_download"),
        ],
        [InlineKeyboardButton("📤 Загрузить", callback_data="schedule_upload")],
    ]
    if is_admin_flag:
        buttons.append(
            [InlineKeyboardButton("👥 Согласующие", callback_data="schedule_approvers")]
        )

    approvers = get_current_approvers(settings)
    if user_tag and user_tag in approvers:
        buttons.append(
            [
                InlineKeyboardButton(
                    f"✅ Согласовать ({user_tag})",
                    callback_data=f"schedule_approve:{user_tag}",
                ),
                InlineKeyboardButton(
                    f"✏️ На доработку ({user_tag})",
                    callback_data=f"schedule_rework:{user_tag}",
                ),
            ]
        )

    return InlineKeyboardMarkup(buttons)


def remarks_menu_inline() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        [
            [
                InlineKeyboardButton(
                    "🔎 Поиск по номеру дела", callback_data="remarks_search_case"
                )
            ],
            [InlineKeyboardButton("🏗 ОНзС", callback_data="remarks_onzs")],
            [InlineKeyboardButton("📥 Открыть файл", callback_data="remarks_download")],
        ]
    )


def inspector_menu_inline() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        [
            [InlineKeyboardButton("➕ Добавить выезд", callback_data="inspector_add")],
            [
                InlineKeyboardButton("📋 Список выездов", callback_data="inspector_list"),
                InlineKeyboardButton(
                    "📥 Скачать Excel", callback_data="inspector_download"
                ),
            ],
            [
                InlineKeyboardButton(
                    "🔄 Обновить", callback_data="inspector_reset"
                )
            ],
        ]
    )


def final_checks_menu_inline() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        [
            [
                InlineKeyboardButton("📅 За неделю", callback_data="final_week"),
                InlineKeyboardButton("📆 За месяц", callback_data="final_month"),
            ],
            [
                InlineKeyboardButton(
                    "📊 Выбрать период", callback_data="final_period"
                )
            ],
            [
                InlineKeyboardButton(
                    "🔎 По номеру дела", callback_data="final_search_case"
                )
            ],
        ]
    )


# -------------------------------------------------
# График
# -------------------------------------------------
def get_schedule_df() -> Optional[pd.DataFrame]:
    SHEET = "График"
    url = build_export_url(GSHEETS_SPREADSHEET_ID)

    try:
        resp = requests.get(url, timeout=30)
        resp.raise_for_status()
    except Exception as e:
        log.error("Ошибка скачивания Excel (график): %s", e)
        return None

    try:
        xls = pd.ExcelFile(BytesIO(resp.content))
        if SHEET not in xls.sheet_names:
            log.error("В файле нет листа '%s'", SHEET)
            return None
        df = pd.read_excel(xls, sheet_name=SHEET)
        df = df.dropna(how="all").reset_index(drop=True)
        return df
    except Exception as e:
        log.error("Ошибка чтения листа графика: %s", e)
        return None


HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


async def send_schedule_xlsx(
    chat_id: int, dataframe: pd.DataFrame, context: ContextTypes.DEFAULT_TYPE
):
    df = dataframe.copy().reset_index(drop=True)
    headers = list(df.columns)

    date_col_name: Optional[str] = None
    for h in headers:
        if "дата выезда" in str(h).lower():
            date_col_name = h
            break
    if date_col_name:
        try:
            df[date_col_name] = pd.to_datetime(
                df[date_col_name], errors="coerce", dayfirst=True
            )
        except Exception:
            pass

    settings = get_schedule_state()
    version = get_schedule_version(settings)
    approvals = get_schedule_approvals(version)

    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        df.to_excel(
            writer,
            sheet_name="График выездов",
            index=False,
            startrow=2,
            header=False,
        )

        wb = writer.book
        ws = writer.sheets["График выездов"]

        for col_num, value in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num, value=value)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for column in ws.columns:
            max_length = 0
            col_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value is not None and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 4, 50)

        ws.freeze_panes = ws["A3"]

        last_col_letter = ws.cell(row=2, column=len(headers)).column_letter
        ws.auto_filter.ref = f"A2:{last_col_letter}{len(df) + 2}"

        for row in ws[f"A3:{last_col_letter}{len(df) + 2}"]:
            for cell in row:
                cell.border = BORDER

        LIGHT_FILL = PatternFill(
            start_color="F0F0F0", end_color="F0F0F0", fill_type="solid"
        )
        for idx, row in enumerate(
            ws.iter_rows(min_row=3, max_row=len(df) + 2), start=3
        ):
            if idx % 2 == 0:
                for cell in row:
                    cell.fill = LIGHT_FILL

        tab = Table(
            displayName="ScheduleTable",
            ref=f"A2:{last_col_letter}{len(df) + 2}",
        )
        tab.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(tab)

        date_idx = None
        onzs_idx = None
        dev_idx = None
        obj_idx = None

        for i, h in enumerate(headers, start=1):
            h_low = str(h).lower()
            if date_idx is None and "дата выезда" in h_low:
                date_idx = i
            if onzs_idx is None and "онзс" in h_low:
                onzs_idx = i
            if dev_idx is None and "наименование застройщика" in h_low:
                dev_idx = i
            if obj_idx is None and "наименование объекта" in h_low:
                obj_idx = i

        for row_idx in range(3, len(df) + 3):
            if date_idx:
                cell = ws.cell(row=row_idx, column=date_idx)
                cell.number_format = "DD.MM.YYYY"
            if onzs_idx:
                cell = ws.cell(row=row_idx, column=onzs_idx)
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=False
                )
            if dev_idx:
                cell = ws.cell(row=row_idx, column=dev_idx)
                cell.alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=True
                )
            if obj_idx:
                cell = ws.cell(row=row_idx, column=obj_idx)
                cell.alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=True
                )

        if approvals:
            last_data_row = len(df) + 2
            summary_start = last_data_row + 2

            header = build_schedule_header(version, approvals)
            ws.merge_cells(f"A{summary_start}:{last_col_letter}{summary_start}")
            cell_header = ws[f"A{summary_start}"]
            cell_header.value = header
            cell_header.font = Font(bold=True, size=12, color="FFFFFF")
            cell_header.fill = PatternFill(
                start_color="4F81BD", end_color="4F81BD", fill_type="solid"
            )
            cell_header.alignment = Alignment(horizontal="center", vertical="center")

            sub_row = summary_start + 1
            ws.merge_cells(f"A{sub_row}:{last_col_letter}{sub_row}")
            cell_sub = ws[f"A{sub_row}"]
            cell_sub.value = "Согласовано всеми:"
            cell_sub.font = Font(bold=True, size=11)
            cell_sub.alignment = Alignment(horizontal="left", vertical="center")

            row_ptr = sub_row + 1
            approved_rows = [r for r in approvals if r["status"] == "approved"]
            others = [r for r in approvals if r["status"] != "approved"]

            list_fill = PatternFill(
                start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
            )

            for r in approved_rows:
                line = f"• {r['approver']} — {_format_dt(r['decided_at'])} ✅"
                ws.merge_cells(f"A{row_ptr}:{last_col_letter}{row_ptr}")
                cell = ws[f"A{row_ptr}"]
                cell.value = line
                cell.fill = list_fill
                cell.font = Font(size=11)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                for col_idx in range(1, len(headers) + 1):
                    ws.cell(row=row_ptr, column=col_idx).border = BORDER
                row_ptr += 1

            if others:
                ws.merge_cells(f"A{row_ptr}:{last_col_letter}{row_ptr}")
                cell_pending = ws[f"A{row_ptr}"]
                cell_pending.value = "⚠ Есть несогласованные/на доработке."
                cell_pending.font = Font(italic=True, color="C00000")
                cell_pending.alignment = Alignment(
                    horizontal="left", vertical="center"
                )
                for col_idx in range(1, len(headers) + 1):
                    ws.cell(row=row_ptr, column=col_idx).border = BORDER

    bio.seek(0)
    filename = f"График_выездов_СОТ_{date.today().strftime('%d.%m.%Y')}.xlsx"

    await context.bot.send_document(
        chat_id=chat_id,
        document=InputFile(bio, filename=filename),
        caption="График выездов отдела СОТ",
    )


# -------------------------------------------------
# Текст графика
# -------------------------------------------------
def _format_dt(iso_str: Optional[str]) -> str:
    if not iso_str:
        return ""
    try:
        dt = datetime.fromisoformat(iso_str)
        return dt.strftime("%d.%m.%Y %H:%M")
    except Exception:
        return iso_str


def _compute_schedule_dates(
    approvals: List[sqlite3.Row],
) -> (Optional[date], Optional[date]):
    dates: List[date] = []
    for r in approvals:
        if r["status"] == "approved" and r["decided_at"]:
            try:
                dt = datetime.fromisoformat(r["decided_at"])
                dates.append(dt.date())
            except Exception:
                pass
    if not dates:
        return None, None
    base = max(dates)
    d_from = base
    d_to = base + timedelta(days=4)
    return d_from, d_to


def build_schedule_header(version: int, approvals: List[sqlite3.Row]) -> str:
    d_from, d_to = _compute_schedule_dates(approvals)
    if not d_from or not d_to:
        return f"📅 График выездов (версия {version})"
    return f"📅 График выездов с {d_from:%d.%m.%Y} по {d_to:%d.%m.%Y} г"


def write_schedule_summary_to_sheet(version: int, approvals: List[sqlite3.Row]) -> None:
    service = get_sheets_service()
    if service is None:
        log.error(
            "Google Sheets сервис недоступен – не могу записать итог согласования в 'График'."
        )
    else:
        sheet_name = "График"
        header = build_schedule_header(version, approvals)
        rows = [
            [""],
            [header],
            ["Согласовано всеми:"],
        ]
        for r in approvals:
            rows.append(
                [f"{r['approver']} — {_format_dt(r['decided_at'])} ✅"]
            )

        body = {"values": rows}

        try:
            service.spreadsheets().values().append(
                spreadsheetId=GSHEETS_SPREADSHEET_ID,
                range=f"'{sheet_name}'!A1",
                valueInputOption="USER_ENTERED",
                insertDataOption="INSERT_ROWS",
                body=body,
            ).execute()
            log.info(
                "Итог согласования версии %s дописан в лист '%s'.",
                version,
                sheet_name,
            )
        except Exception as e:
            log.error(
                "Ошибка записи итога согласования в лист '%s': %s", sheet_name, e
            )


def build_schedule_text(is_admin_flag: bool, settings: dict) -> str:
    version = get_schedule_version(settings)
    approvals = get_schedule_approvals(version)
    approvers = get_current_approvers(settings)

    header = build_schedule_header(version, approvals)
    lines = [header, ""]

    if not approvers:
        lines.append("Согласующие не назначены.")
        return "\n".join(lines)

    pending: List[str] = []
    approved: List[sqlite3.Row] = []
    rework: List[sqlite3.Row] = []

    by_approver = {r["approver"]: r for r in approvals}

    for a in approvers:
        r = by_approver.get(a)
        if not r or r["status"] == "pending":
            pending.append(a)
        elif r["status"] == "approved":
            approved.append(r)
        elif r["status"] == "rework":
            rework.append(r)

    if rework:
        lines.append("Отправлено на доработку:")
        for r in rework:
            lines.append(
                f"• {r['approver']} — {_format_dt(r['decided_at'])} (Комментарий: {r['comment'] or 'нет'})"
            )
    elif pending:
        lines.append("На согласовании у:")
        for a in pending:
            lines.append(
                f"• {a} — запрошено {_format_dt(by_approver[a]['requested_at'])}"
            )
        if approved:
            lines.append("")
            lines.append("Уже согласовали:")
            for r in approved:
                lines.append(f"• {r['approver']} — {_format_dt(r['decided_at'])} ✅")
    else:
        lines.append("Согласовано всеми:")
        for r in approved:
            lines.append(f"• {r['approver']} — {_format_dt(r['decided_at'])} ✅")

    return "\n".join(lines)


# -------------------------------------------------
# Замечания: НЕ УСТРАНЕНЫ
# -------------------------------------------------
def build_remarks_not_done_text(df: pd.DataFrame) -> str:
    COLS = {
        "case": "I",
        "pb": "Q",
        "pb_zk": "R",
        "ar": "X",
        "eom": "AD",
    }

    TITLES = {
        "pb": "Отметка об устранении замечаний ПБ да/нет",
        "pb_zk": "Отметка об устранении замечаний ПБ в ЗК КНД да/нет",
        "ar": "Отметка об устранении нарушений АР, ММГН, АГО да/нет",
        "eom": "Отметка об устранении нарушений ЭОМ да/нет",
    }

    idx_case = excel_col_to_index(COLS["case"])
    idx_pb = excel_col_to_index(COLS["pb"])
    idx_pb_zk = excel_col_to_index(COLS["pb_zk"])
    idx_ar = excel_col_to_index(COLS["ar"])
    idx_eom = excel_col_to_index(COLS["eom"])

    def is_net(val):
        if val is None:
            return False
        text = str(val).lower().replace("\n", " ").strip()
        if not text or text in {"-", "н/д"}:
            return False
        return text.startswith("нет")

    grouped = {}

    for _, row in df.iterrows():
        case = str(row.iloc[idx_case]).strip()
        if not case:
            continue

        flags = {
            "pb": is_net(row.iloc[idx_pb]) if idx_pb < len(row) else False,
            "pb_zk": is_net(row.iloc[idx_pb_zk]) if idx_pb_zk < len(row) else False,
            "ar": is_net(row.iloc[idx_ar]) if idx_ar < len(row) else False,
            "eom": is_net(row.iloc[idx_eom]) if idx_eom < len(row) else False,
        }

        if not any(flags.values()):
            continue

        if case not in grouped:
            grouped[case] = {"pb": set(), "ar": set(), "eom": set()}

        if flags["pb"]:
            grouped[case]["pb"].add(TITLES["pb"])
        if flags["pb_zk"]:
            grouped[case]["pb"].add(TITLES["pb_zk"])
        if flags["ar"]:
            grouped[case]["ar"].add(TITLES["ar"])
        if flags["eom"]:
            grouped[case]["eom"].add(TITLES["eom"])

    if not grouped:
        return "Во всех строках нет статусов «нет»."

    lines = [
        "Строки со статусом «НЕ УСТРАНЕНЫ (нет)»",
        "",
        "Лист: " + get_current_remarks_sheet_name(),
        "",
    ]

    for case, blocks in grouped.items():
        parts = []
        if blocks["pb"]:
            parts.append(
                "Пожарная безопасность: "
                + ", ".join(b + " - нет" for b in blocks["pb"])
            )
        if blocks["ar"]:
            parts.append(
                "Архитектура, ММГН, АГО: "
                + ", ".join(b + " - нет" for b in blocks["ar"])
            )
        if blocks["eom"]:
            parts.append(
                "Электроснабжение: "
                + ", ".join(b + " - нет" for b in blocks["eom"])
            )
        lines.append(f"• {case} — " + "; ".join(parts))

    return "\n".join(lines)


def build_remarks_not_done_by_onzs(df: pd.DataFrame, onzs_value: str) -> str:
    sheet_name = get_current_remarks_sheet_name()

    onzs_idx = get_col_index_by_header(df, "онзс", "D")
    if onzs_idx is None:
        return "Не удалось определить столбец ОНзС в файле замечаний."

    COLS = {
        "case": "I",
        "pb": "Q",
        "pb_zk": "R",
        "ar": "X",
        "eom": "AD",
    }

    TITLES = {
        "pb": "Отметка об устранении замечаний ПБ да/нет",
        "pb_zk": "Отметка об устранении замечаний ПБ в ЗК КНД да/нет",
        "ar": "Отметка об устранении нарушений АР, ММГН, АГО да/нет",
        "eom": "Отметка об устранении нарушений ЭОМ да/нет",
    }

    idx_case = excel_col_to_index(COLS["case"])
    idx_pb = excel_col_to_index(COLS["pb"])
    idx_pb_zk = excel_col_to_index(COLS["pb_zk"])
    idx_ar = excel_col_to_index(COLS["ar"])
    idx_eom = excel_col_to_index(COLS["eom"])

    def is_net(val):
        if val is None:
            return False
        text = str(val).lower().replace("\n", " ").strip()
        if not text or text in {"-", "н/д"}:
            return False
        return text.startswith("нет")

    grouped = {}

    num_str = normalize_onzs_value(onzs_value)

    for _, row in df.iterrows():
        try:
            val_raw = row.iloc[onzs_idx]
        except Exception:
            val_raw = None

        val_norm = normalize_onzs_value(val_raw)
        if val_norm != num_str:
            continue

        case = ""
        try:
            case = str(row.iloc[idx_case]).strip()
        except Exception:
            pass

        if not case:
            continue

        flags = {
            "pb": is_net(row.iloc[idx_pb]) if idx_pb < len(row) else False,
            "pb_zk": is_net(row.iloc[idx_pb_zk]) if idx_pb_zk < len(row) else False,
            "ar": is_net(row.iloc[idx_ar]) if idx_ar < len(row) else False,
            "eom": is_net(row.iloc[idx_eom]) if idx_eom < len(row) else False,
        }

        if not any(flags.values()):
            continue

        if case not in grouped:
            grouped[case] = {"pb": set(), "ar": set(), "eom": set()}

        if flags["pb"]:
            grouped[case]["pb"].add(TITLES["pb"])
        if flags["pb_zk"]:
            grouped[case]["pb"].add(TITLES["pb_zk"])
        if flags["ar"]:
            grouped[case]["ar"].add(TITLES["ar"])
        if flags["eom"]:
            grouped[case]["eom"].add(TITLES["eom"])

    if not grouped:
        return (
            f"По ОНзС {onzs_value} нет строк со статусом «нет».\n"
            f"Лист: {sheet_name}"
        )

    lines = [
        f"Строки со статусом «НЕ УСТРАНЕНЫ (нет)» по ОНзС {onzs_value}",
        "",
        "Лист: " + sheet_name,
        "",
    ]

    for case, blocks in grouped.items():
        parts = []
        if blocks["pb"]:
            parts.append(
                "Пожарная безопасность: "
                + ", ".join(b + " - нет" for b in blocks["pb"])
            )
        if blocks["ar"]:
            parts.append(
                "Архитектура, ММГН, АГО: "
                + ", ".join(b + " - нет" for b in blocks["ar"])
            )
        if blocks["eom"]:
            parts.append(
                "Электроснабжение: "
                + ", ".join(b + " - нет" for b in blocks["eom"])
            )
        lines.append(f"• {case} — " + "; ".join(parts))

    return "\n".join(lines)


def build_case_cards_text(df: pd.DataFrame, case_no: str) -> str:
    sheet_name = get_current_remarks_sheet_name()

    case_no = case_no.strip()
    if not case_no:
        return "Номер дела не указан."

    target = normalize_case_number(case_no)

    idx_case = get_case_col_index(df)
    if idx_case is None:
        return (
            "Не удалось определить столбец «Номер дела (I)» в файле замечаний. "
            "Проверьте структуру листа."
        )

    idx_date = get_col_index_by_header(df, "дата выезда", "B")
    idx_onzs = get_col_index_by_header(df, "онзс", "D")
    idx_dev = get_col_index_by_header(df, "наименование застройщика", "F")
    idx_obj = get_col_index_by_header(df, "наименование объекта", "G")
    idx_addr = get_col_index_by_header(df, "строительный адрес", "H")

    idx_pb = excel_col_to_index("Q")
    idx_pb_zk = excel_col_to_index("R")
    idx_ar = excel_col_to_index("X")
    idx_eom = excel_col_to_index("AD")

    mask: List[bool] = []
    for _, row in df.iterrows():
        try:
            val_raw = row.iloc[idx_case]
        except Exception:
            val_raw = None
        val_norm = normalize_case_number(val_raw)
        mask.append(val_norm == target)

    if not any(mask):
        return (
            f"По номеру дела {case_no} ничего не найдено.\n"
            f"Лист: {sheet_name}"
        )

    df_sel = df[mask]

    lines: List[str] = [
        f"Результаты поиска по номеру дела: {case_no}",
        "",
        f"Лист: {sheet_name}",
        "",
    ]

    for _, row in df_sel.iterrows():

        def safe(idx: Optional[int]) -> str:
            if idx is None:
                return ""
            try:
                return str(row.iloc[idx]).strip()
            except Exception:
                return ""

        date_raw = safe(idx_date)
        date_fmt = date_raw
        try:
            if date_raw:
                dt = pd.to_datetime(date_raw, dayfirst=True, errors="ignore")
                if isinstance(dt, (datetime, pd.Timestamp)):
                    date_fmt = dt.strftime("%d.%m.%Y")
        except Exception:
            pass

        onzs_val = safe(idx_onzs)
        dev_val = safe(idx_dev)
        obj_val = safe(idx_obj)
        addr_val = safe(idx_addr)

        def safe_status(idx: int) -> str:
            try:
                if idx < len(row):
                    return str(row.iloc[idx]).strip()
            except Exception:
                pass
            return ""

        pb_val = safe_status(idx_pb)
        pb_zk_val = safe_status(idx_pb_zk)
        ar_val = safe_status(idx_ar)
        eom_val = safe_status(idx_eom)

        lines.append(f"Номер дела: {case_no}")
        if date_fmt:
            lines.append(f"Дата выезда: {date_fmt}")
        if onzs_val:
            lines.append(f"ОНзС: {onzs_val}")
        if dev_val:
            lines.append(f"Застройщик: {dev_val}")
        if obj_val:
            lines.append(f"Объект: {obj_val}")
        if addr_val:
            lines.append(f"Адрес: {addr_val}")

        lines.append("")
        lines.append(f"ПБ: {pb_val or '-'}")
        lines.append(f"ПБ ЗК: {pb_zk_val or '-'}")
        lines.append(f"АР/ММГН/АГО: {ar_val or '-'}")
        lines.append(f"ЭОМ: {eom_val or '-'}")
        lines.append("")
        lines.append("────────────")
        lines.append("")

    return "\n".join(lines)


# -------------------------------------------------
# Отправка длинного текста
# -------------------------------------------------
async def send_long_text(chat, text: str, chunk_size=3500):
    lines = text.split("\n")
    buf = ""

    for line in lines:
        if len(buf) + len(line) + 1 > chunk_size:
            await chat.send_message(buf)
            buf = line
        else:
            buf = buf + "\n" + line if buf else line

    if buf:
        await chat.send_message(buf)


# -------------------------------------------------
# Лист замечаний
# -------------------------------------------------
def get_remarks_df_current() -> Optional[pd.DataFrame]:
    sheet = get_current_remarks_sheet_name()
    url = build_export_url(GSHEETS_SPREADSHEET_ID)

    try:
        resp = requests.get(url, timeout=30)
        resp.raise_for_status()
        xls = pd.ExcelFile(BytesIO(resp.content))
        if sheet not in xls.sheet_names:
            log.error("В файле нет листа '%s'", sheet)
            return None
        return pd.read_excel(xls, sheet_name=sheet)
    except Exception as e:
        log.error("Ошибка чтения листа замечаний: %s", e)
        return None


# -------------------------------------------------
# Итоговые проверки: чтение, фильтр, текст, Excel
# -------------------------------------------------
def get_final_checks_df() -> Optional[pd.DataFrame]:
    """
    Читает файл итоговых проверок из отдельной таблицы FINAL_CHECKS_SPREADSHEET_ID.
    Берём первый лист книги.
    """
    sheet_id = FINAL_CHECKS_SPREADSHEET_ID
    if not sheet_id:
        log.error("FINAL_CHECKS_SPREADSHEET_ID не задан.")
        return None

    url = build_export_url(sheet_id)

    try:
        resp = requests.get(url, timeout=30)
        resp.raise_for_status()
    except Exception as e:
        log.error("Ошибка скачивания Excel (итоговые проверки): %s", e)
        return None

    try:
        xls = pd.ExcelFile(BytesIO(resp.content))
        if not xls.sheet_names:
            log.error("Файл итоговых проверок пуст (нет листов).")
            return None
        sheet_name = xls.sheet_names[0]
        df = pd.read_excel(xls, sheet_name=sheet_name)
        df = df.dropna(how="all").reset_index(drop=True)
        return df
    except Exception as e:
        log.error("Ошибка чтения листа итоговых проверок: %s", e)
        return None


def _parse_final_date(val) -> Optional[date]:
    """
    Преобразует значение из столбцов O/P в дату.
    Поддерживает текстовые и «экселевские» даты.
    """
    if val is None:
        return None
    try:
        if isinstance(val, (datetime, pd.Timestamp)):
            return val.date()
        if isinstance(val, (int, float)) and not pd.isna(val):
            dt = pd.to_datetime(val, errors="coerce")
            if isinstance(dt, (datetime, pd.Timestamp)):
                return dt.date()
        dt = pd.to_datetime(str(val), dayfirst=True, errors="coerce")
        if isinstance(dt, (datetime, pd.Timestamp)):
            return dt.date()
    except Exception:
        return None
    return None


def filter_final_checks_df(
    df: pd.DataFrame,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    case_no: Optional[str] = None,
    basis: str = "any",  # "start" -> только O, "end" -> только P, "any" -> O или P
) -> pd.DataFrame:
    idx_case = excel_col_to_index("B")
    idx_start = excel_col_to_index("O")
    idx_end = excel_col_to_index("P")

    basis = (basis or "any").lower()
    if basis not in ("start", "end", "any"):
        basis = "any"

    case_filter_norm = normalize_case_number(case_no) if case_no else None

    mask: List[bool] = []
    for _, row in df.iterrows():
        include = True

        # --- фильтр по номеру дела ---
        if case_filter_norm:
            try:
                case_val = row.iloc[idx_case]
            except Exception:
                case_val = None
            val_norm = normalize_case_number(case_val)
            if not val_norm or val_norm != case_filter_norm:
                include = False

        # --- фильтр по периоду ---
        if include and start_date and end_date:
            try:
                s_raw = row.iloc[idx_start]
            except Exception:
                s_raw = None
            try:
                e_raw = row.iloc[idx_end]
            except Exception:
                e_raw = None

            d_start = _parse_final_date(s_raw)
            d_end = _parse_final_date(e_raw)

            if basis == "start":
                base = d_start
            elif basis == "end":
                base = d_end
            else:  # "any"
                base = d_start or d_end

            if base is None or base < start_date or base > end_date:
                include = False

        mask.append(include)

    if not mask:
        return df.iloc[0:0].copy()

    df_f = df[mask].copy().reset_index(drop=True)
    return df_f


def build_final_checks_text_filtered(
    df: pd.DataFrame,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    case_no: Optional[str] = None,
    header: str = "📋 Итоговые проверки",
    basis: str = "any",  # "start" / "end" / "any"
) -> str:
    """
    Универсальный вывод итоговых проверок:
    - фильтр по периоду (start_date / end_date) по выбранной базе (O или P);
    - фильтр по номеру дела (case_no).
    """
    df_f = filter_final_checks_df(
        df,
        start_date=start_date,
        end_date=end_date,
        case_no=case_no,
        basis=basis,
    )

    idx_case = excel_col_to_index("B")
    idx_obj = excel_col_to_index("D")
    idx_addr = excel_col_to_index("E")
    idx_start = excel_col_to_index("O")
    idx_end = excel_col_to_index("P")

    lines: List[str] = [header, ""]

    if df_f.empty:
        if case_no:
            return (
                f"По номеру дела {case_no} в таблице итоговых проверок ничего не найдено."
            )
        if start_date and end_date:
            return (
                f"За период {start_date:%d.%m.%Y} — {end_date:%d.%m.%Y} "
                f"итоговые проверки не найдены."
            )
        return "В таблице итоговых проверок нет строк с заполненным номером дела (B)."

    for _, row in df_f.iterrows():

        def safe_text(idx: int) -> str:
            try:
                val = row.iloc[idx]
            except Exception:
                return ""
            if pd.isna(val):
                return ""
            return str(val).strip()

        case_val = safe_text(idx_case)
        if not case_val:
            continue

        obj = safe_text(idx_obj)
        addr = safe_text(idx_addr)

        d_start_raw = row.iloc[idx_start] if idx_start < len(row) else None
        d_end_raw = row.iloc[idx_end] if idx_end < len(row) else None

        row_start = _parse_final_date(d_start_raw)
        row_end = _parse_final_date(d_end_raw)

        def fmt_date(d: Optional[date]) -> str:
            return d.strftime("%d.%m.%Y") if d else ""

        d_start = fmt_date(row_start)
        d_end = fmt_date(row_end)

        lines.append(f"Номер дела: {case_val}")
        if obj:
            lines.append(f"Объект: {obj}")
        if addr:
            lines.append(f"Адрес: {addr}")
        if d_start or d_end:
            if d_start and d_end:
                lines.append(f"Период итоговой проверки: {d_start} — {d_end}")
            elif d_start:
                lines.append(f"Дата начала итоговой проверки: {d_start}")
            else:
                lines.append(f"Дата окончания итоговой проверки: {d_end}")
        lines.append("")
        lines.append("────────────")
        lines.append("")

    return "\n".join(lines)


def build_final_checks_text(df: pd.DataFrame) -> str:
    """
    Старый интерфейс (без фильтров) — на всякий случай.
    """
    return build_final_checks_text_filtered(df)


async def send_final_checks_xlsx_filtered(
    chat_id: int,
    df: pd.DataFrame,
    context: ContextTypes.DEFAULT_TYPE,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    case_no: Optional[str] = None,
    filename_suffix: str = "",
    basis: str = "any",
):
    df_f = filter_final_checks_df(
        df,
        start_date=start_date,
        end_date=end_date,
        case_no=case_no,
        basis=basis,
    )
    if df_f.empty:
        await context.bot.send_message(
            chat_id=chat_id,
            text="Нет данных для выгрузки по выбранным условиям.",
        )
        return

    bio = BytesIO()
    df_f.to_excel(bio, sheet_name="Итоговые проверки", index=False)
    bio.seek(0)

    fname = "Итоговые_проверки"
    parts = []
    if case_no:
        parts.append(f"дело_{case_no}")
    if start_date and end_date:
        parts.append(f"{start_date:%d.%m.%Y}-{end_date:%d.%m.%Y}")
    if filename_suffix:
        parts.append(filename_suffix)
    if parts:
        fname += "_" + "_".join(parts)
    fname += ".xlsx"

    await context.bot.send_document(
        chat_id=chat_id,
        document=InputFile(bio, filename=fname),
        caption="Итоговые проверки (фильтрованный список)",
    )


# -------------------------------------------------
# Инспектор → Google Sheets
# -------------------------------------------------
def append_inspector_row_to_excel(form: Dict[str, Any]) -> bool:
    service = get_sheets_service()
    if service is None:
        log.error("Google Sheets API недоступен.")
        return False

    try:
        area_str = str(form.get("area", "")).replace(".", ",")
        floors_str = str(form.get("floors", ""))

        d_value = (
            f"Площадь (кв.м): {area_str}\n"
            f"Количество этажей: {floors_str}"
        )

        row = [
            "",
            form.get("date").strftime("%d.%m.%Y") if form.get("date") else "",
            "",
            d_value,
            form.get("onzs", ""),
            form.get("developer", ""),
            form.get("object", ""),
            form.get("address", ""),
            form.get("case", ""),
            form.get("check_type", ""),
        ]

        body = {"values": [row]}

        response = (
            service.spreadsheets()
            .values()
            .append(
                spreadsheetId=GSHEETS_SPREADSHEET_ID,
                range=f"'{INSPECTOR_SHEET_NAME}'!A1",
                valueInputOption="USER_ENTERED",
                insertDataOption="INSERT_ROWS",
                body=body,
            )
            .execute()
        )

        log.info("Инспектор: запись добавлена в Google Sheets: %s", response)
        return True

    except Exception as e:
        log.error("Ошибка записи инспектора в Google Sheets: %s", e)
        return False


# -------------------------------------------------
# Инспектор — мастер
# -------------------------------------------------
async def inspector_process(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    form = context.user_data.get("inspector_form", {}) or {}
    step = form.get("step")

    if not step:
        context.user_data["inspector_form"] = {"step": "date"}
        await update.message.reply_text(
            "👮‍♂️ Выезд инспектора\n\n"
            "1/8. Дата выезда (ДД.ММ.ГГГГ):"
        )
        return

    if step == "date":
        try:
            form["date"] = datetime.strptime(text, "%d.%m.%Y").date()
            form["step"] = "area"
            context.user_data["inspector_form"] = form
            await update.message.reply_text("1/8. Площадь объекта (кв.м):")
        except Exception:
            await update.message.reply_text(
                "Введите дату в формате ДД.ММ.ГГГГ (например, 30.12.2025)"
            )
        return

    if step == "area":
        form["area"] = text
        form["step"] = "floors"
        context.user_data["inspector_form"] = form
        await update.message.reply_text("2/8. Количество этажей:")
        return

    if step == "floors":
        form["floors"] = text
        form["step"] = "onzs"
        context.user_data["inspector_form"] = form
        await update.message.reply_text("3/8. ОНзС (1–12):")
        return

    if step == "onzs":
        form["onzs"] = text
        form["step"] = "developer"
        context.user_data["inspector_form"] = form
        await update.message.reply_text("4/8. Наименование застройщика:")
        return

    if step == "developer":
        form["developer"] = text
        form["step"] = "object"
        context.user_data["inspector_form"] = form
        await update.message.reply_text("5/8. Наименование объекта:")
        return

    if step == "object":
        form["object"] = text
        form["step"] = "address"
        context.user_data["inspector_form"] = form
        await update.message.reply_text("6/8. Строительный адрес:")
        return

    if step == "address":
        form["address"] = text
        form["step"] = "case"
        context.user_data["inspector_form"] = form
        await update.message.reply_text("7/8. Номер дела (формат 00-00-000000):")
        return

    if step == "case":
        form["case"] = text
        form["step"] = "check_type"
        context.user_data["inspector_form"] = form
        await update.message.reply_text(
            "8/8. Вид проверки (ПП, итоговая, профвизит, поручение и т.п.):"
        )
        return

    if step == "check_type":
        form["check_type"] = text
        form["step"] = "done"
        context.user_data["inspector_form"] = form

        await update.message.reply_text("⏳ Сохраняю выезд...")

        ok_db = save_inspector_to_db(form)
        ok_gs = append_inspector_row_to_excel(form)

        if ok_db and ok_gs:
            msg = "✅ Выезд сохранён в боте и добавлен в общую таблицу."
        elif ok_db and not ok_gs:
            msg = (
                "✅ Выезд сохранён в боте.\n"
                "⚠ Не удалось добавить в Google Sheets (проверьте ключ/права)."
            )
        elif not ok_db and ok_gs:
            msg = (
                "⚠ Выезд добавлен в Google Sheets, но не удалось сохранить локную запись."
            )
        else:
            msg = (
                "❌ Не удалось сохранить выезд ни локно, ни в Google Sheets.\n"
                "Сообщите разработчику."
            )

        await update.message.reply_text(msg)
        context.user_data.pop("inspector_form", None)
        return


# -------------------------------------------------
# ОНзС
# -------------------------------------------------
def onzs_menu_inline() -> InlineKeyboardMarkup:
    buttons = []
    row = []
    for i in range(1, 13):
        row.append(InlineKeyboardButton(str(i), callback_data=f"onzs_filter_{i}"))
        if len(row) == 4:
            buttons.append(row)
            row = []
    if row:
        buttons.append(row)
    return InlineKeyboardMarkup(buttons)


def build_onzs_list_by_number(df: pd.DataFrame, number: str) -> str:
    onzs_idx = get_col_index_by_header(df, "онзс", "D")
    if onzs_idx is None:
        return "Не удалось определить столбец ОНзС в файле замечаний."

    case_idx = get_case_col_index(df)
    addr_idx = get_col_index_by_header(df, "строительный адрес", "H")

    num_str = normalize_onzs_value(number)
    mask: List[bool] = []
    for _, row in df.iterrows():
        try:
            val_raw = row.iloc[onzs_idx]
        except Exception:
            val_raw = None
        val_norm = normalize_onzs_value(val_raw)
        mask.append(val_norm == num_str)

    if not any(mask):
        return f"Нет объектов с ОНзС = {number}."

    df_f = df[mask]

    lines = [f"ОНзС = {number}", f"Найдено дел: {len(df_f)}", ""]

    for _, row in df_f.iterrows():

        def safe(idx: Optional[int]) -> str:
            if idx is None:
                return ""
            try:
                val = row.iloc[idx]
            except Exception:
                return ""
            try:
                if pd.isna(val):
                    return ""
            except Exception:
                pass
            s = str(val).strip()
            if not s or s.lower() == "nan":
                return ""
            return s

        case_no = safe(case_idx)
        addr = safe(addr_idx)

        if not case_no and not addr:
            continue

        if case_no and addr:
            lines.append(f"• {case_no} — {addr}")
        elif case_no:
            lines.append(f"• {case_no}")
        else:
            lines.append(f"• {addr}")

    return "\n".join(lines)


# -------------------------------------------------
# Инспектор — список/Excel
# -------------------------------------------------
def build_inspector_list_text(rows: List[sqlite3.Row]) -> str:
    if not rows:
        return "Пока нет сохранённых выездов инспектора."

    lines: List[str] = ["Последние выезды инспектора:", ""]
    for r in rows:
        d = r["date"] or ""
        try:
            d_fmt = datetime.strptime(d, "%Y-%m-%d").strftime("%d.%m.%Y")
        except Exception:
            d_fmt = d
        lines.append(
            f"• {d_fmt} — дело {r['case_no'] or '-'}, "
            f"ОНзС {r['onzs'] or '-'}, {r['check_type'] or ''}"
        )
        addr = r["address"] or ""
        if addr:
            lines.append(f"  Адрес: {addr}")
        obj = r["object"] or ""
        if obj:
            lines.append(f"  Объект: {obj}")
        dev = r["developer"] or ""
        if dev:
            lines.append(f"  Застройщик: {dev}")
        lines.append("")
    return "\n".join(lines)


async def send_inspector_xlsx(
    chat_id: int, rows: List[sqlite3.Row], context: ContextTypes.DEFAULT_TYPE
):
    if not rows:
        await context.bot.send_message(
            chat_id=chat_id, text="Пока нет сохранённых выездов инспектора."
        )
        return

    data = []
    for r in rows:
        d = r["date"] or ""
        try:
            d_fmt = datetime.strptime(d, "%Y-%m-%d").strftime("%d.%m.%Y")
        except Exception:
            d_fmt = d
        data.append(
            {
                "Дата выезда": d_fmt,
                "Площадь (кв.м)": r["area"] or "",
                "Этажность": r["floors"] or "",
                "ОНзС": r["onzs"] or "",
                "Застройщик": r["developer"] or "",
                "Наименование объекта": r["object"] or "",
                "Строительный адрес": r["address"] or "",
                "Номер дела": r["case_no"] or "",
                "Вид проверки": r["check_type"] or "",
            }
        )

    df = pd.DataFrame(data)

    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Инспектор", index=False)

    bio.seek(0)
    filename = f"Инспектор_выезды_{date.today().strftime('%d.%m.%Y')}.xlsx"

    await context.bot.send_document(
        chat_id=chat_id,
        document=InputFile(bio, filename=filename),
        caption="Выезды инспектора (отдельный файл)",
    )


# -------------------------------------------------
# CALLBACK HANDLER
# -------------------------------------------------
async def callback_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    data = query.data
    user = query.from_user
    await query.answer()

    settings = get_schedule_state()
    version = get_schedule_version(settings)

    # --- ГРАФИК ---
    if data == "schedule_refresh":
        df = get_schedule_df()
        if df is None:
            await query.message.reply_text("Не удалось прочитать лист «График».")
        else:
            await query.message.reply_text(f"Лист «График» прочитан, строк: {len(df)}.")
        return

    if data == "schedule_download":
        df = get_schedule_df()
        if df is None or df.empty:
            await query.message.reply_text(
                "Не удалось получить лист «График» для выгрузки."
            )
            return

        await send_schedule_xlsx(
            chat_id=query.message.chat.id,
            dataframe=df,
            context=context,
        )
        return

    if data == "schedule_upload":
        await query.message.reply_text("Загрузка графика в этой сборке не реализована.")
        return

    if data == "schedule_approvers":
        if not is_admin(user.id):
            await query.message.reply_text(
                "Только администратор может настраивать согласующих."
            )
            return
        context.user_data["awaiting_approvers_input"] = {"version": version}
        await query.message.reply_text(
            "Отправьте список согласующих (юзернеймы через пробел/запятую/новую строку), например:\n"
            "@asdinamitif @FrolovAlNGSN @cappit_G59"
        )
        return

    if data.startswith("schedule_approve:") or data.startswith("schedule_rework:"):
        action, approver_tag = data.split(":", 1)
        user_username = user.username or ""
        user_tag = f"@{user_username}" if user_username else ""

        if user_tag.lower() != approver_tag.lower():
            await query.answer(
                text=f"Эта кнопка предназначена для {approver_tag}.",
                show_alert=True,
            )
            return

        if action == "schedule_approve":
            update_schedule_approval_status(version, approver_tag, "approved", None)
            await query.message.reply_text(
                f"{approver_tag} согласовал(а) график. Спасибо!"
            )

            approvals = get_schedule_approvals(version)
            if approvals and all(r["status"] == "approved" for r in approvals):
                header = build_schedule_header(version, approvals)
                lines = [header, "", "Согласовано всеми:"]
                for r in approvals:
                    lines.append(
                        f"• {r['approver']} — {_format_dt(r['decided_at'])} ✅"
                    )
                text = "\n".join(lines)

                write_schedule_summary_to_sheet(version, approvals)

                if SCHEDULE_NOTIFY_CHAT_ID is not None:
                    try:
                        await context.bot.send_message(
                            chat_id=SCHEDULE_NOTIFY_CHAT_ID, text=text
                        )
                    except Exception as e:
                        log.error(
                            "Ошибка отправки графика в канал %s: %s",
                            SCHEDULE_NOTIFY_CHAT_ID,
                            e,
                        )
            return

        if action == "schedule_rework":
            context.user_data["awaiting_rework_comment"] = {
                "version": version,
                "approver": approver_tag,
            }
            await query.message.reply_text(
                "Напишите комментарий, почему график нужно доработать."
            )
            return

    # --- ЗАМЕЧАНИЯ ---
    if data == "remarks_search_case":
        context.user_data["awaiting_case_search"] = True
        await query.message.reply_text(
            "Введите номер дела (формат 00-00-000000), который нужно найти:"
        )
        return

    if data == "remarks_onzs":
        kb = onzs_menu_inline()
        msg = (
            "🏗 Раздел «ОНзС»\n\n"
            "Выберите номер ОНзС, чтобы увидеть список дел (Номер дела (I) + адрес) "
            "из текущего файла замечаний.\n"
            "Для выбранного ОНзС можно отдельно показать только неустранённые замечания."
        )
        await query.message.reply_text(msg, reply_markup=kb)
        return

    if data == "remarks_not_done":
        await query.message.reply_text("Ищу строки со статусом «нет»...")
        df = get_remarks_df_current()
        if df is None:
            await query.message.reply_text(
                "Не удалось получить файл замечаний. Проверьте доступ к таблице."
            )
            return
        text = build_remarks_not_done_text(df)
        await send_long_text(query.message.chat, text)
        return

    if data == "remarks_download":
        await query.message.reply_text(
            "Файл с замечаниями и графиком можно открыть по ссылке:\n"
            f"{GOOGLE_SHEET_URL_DEFAULT}"
        )
        return

    if data.startswith("onzs_filter_"):
        number = data.replace("onzs_filter_", "")
        df = get_remarks_df_current()
        if df is None:
            await query.message.reply_text("Не удалось открыть таблицу ОНзС.")
            return
        text = build_onzs_list_by_number(df, number)
        await send_long_text(query.message.chat, text)

        kb = InlineKeyboardMarkup(
            [
                [
                    InlineKeyboardButton(
                        f"❌ Не устранены (ОНзС {number})",
                        callback_data=f"onzs_not_done_{number}",
                    )
                ]
            ]
        )
        await query.message.reply_text(
            f"Для ОНзС {number} можно показать только строки, где статус «нет».",
            reply_markup=kb,
        )
        return

    if data.startswith("onzs_not_done_"):
        number = data.replace("onzs_not_done_", "")
        df = get_remarks_df_current()
        if df is None:
            await query.message.reply_text(
                "Не удалось получить файл замечаний. Проверьте доступ к таблице."
            )
            return
        text = build_remarks_not_done_by_onzs(df, number)
        await send_long_text(query.message.chat, text)
        return

    # --- ИНСПЕКТОР ---
    if data == "inspector_add":
        context.user_data["inspector_form"] = {"step": "date"}
        await query.message.reply_text(
            "👮‍♂️ Выезд инспектора\n\n"
            "Укажем данные по шагам.\n"
            "1/8. Дата выезда (ДД.ММ.ГГГГ):"
        )
        return

    if data == "inspector_list":
        rows = fetch_inspector_visits(limit=50)
        text = build_inspector_list_text(rows)
        await send_long_text(query.message.chat, "\n".join(text.split("\n")))
        return

    if data == "inspector_download":
        rows = fetch_inspector_visits(limit=1000)
        await send_inspector_xlsx(
            chat_id=query.message.chat.id, rows=rows, context=context
        )
        return

    if data == "inspector_reset":
        clear_inspector_visits()
        await query.message.reply_text(
            "Список выездов инспектора очищен.\n"
            "Новые выезды будут попадать в Excel после добавления через кнопку «➕ Добавить выезд»."
        )
        return

    # --- ИТОГОВЫЕ ПРОВЕРКИ ---
    if data == "final_week":
        # запоминаем режим и спрашиваем, по какой дате фильтровать
        context.user_data["final_range_choice"] = {"mode": "week"}
        kb = InlineKeyboardMarkup(
            [
                [
                    InlineKeyboardButton(
                        "📌 По дате начала (O)", callback_data="final_basis_start"
                    ),
                    InlineKeyboardButton(
                        "📌 По дате окончания (P)", callback_data="final_basis_end"
                    ),
                ]
            ]
        )
        await query.message.reply_text(
            "За неделю: по какой дате фильтровать?\n\n"
            "• O — дата начала итоговой проверки\n"
            "• P — дата окончания итоговой проверки",
            reply_markup=kb,
        )
        return

    if data == "final_month":
        context.user_data["final_range_choice"] = {"mode": "month"}
        kb = InlineKeyboardMarkup(
            [
                [
                    InlineKeyboardButton(
                        "📌 По дате начала (O)", callback_data="final_basis_start"
                    ),
                    InlineKeyboardButton(
                        "📌 По дате окончания (P)", callback_data="final_basis_end"
                    ),
                ]
            ]
        )
        await query.message.reply_text(
            "За месяц: по какой дате фильтровать?\n\n"
            "• O — дата начала итоговой проверки\n"
            "• P — дата окончания итоговой проверки",
            reply_markup=kb,
        )
        return

    if data == "final_period":
        context.user_data["final_range_choice"] = {"mode": "period"}
        kb = InlineKeyboardMarkup(
            [
                [
                    InlineKeyboardButton(
                        "📌 По дате начала (O)", callback_data="final_basis_start"
                    ),
                    InlineKeyboardButton(
                        "📌 По дате окончания (P)", callback_data="final_basis_end"
                    ),
                ]
            ]
        )
        await query.message.reply_text(
            "Выбор периода: по какой дате фильтровать?\n\n"
            "• O — дата начала итоговой проверки\n"
            "• P — дата окончания итоговой проверки",
            reply_markup=kb,
        )
        return

    # выбор базы: O или P
    if data in ("final_basis_start", "final_basis_end"):
        basis = "start" if data == "final_basis_start" else "end"
        state = context.user_data.get("final_range_choice")
        if not state:
            await query.message.reply_text(
                "Сначала выберите режим (за неделю/за месяц/выбрать период) в разделе «Итоговые проверки»."
            )
            return

        mode = state.get("mode")
        # недельный и месячный режимы
        if mode in ("week", "month"):
            df = get_final_checks_df()
            if df is None:
                await query.message.reply_text(
                    "Не удалось открыть таблицу итоговых проверок."
                )
                context.user_data.pop("final_range_choice", None)
                return

            today = local_now().date()
            if mode == "week":
                start = today - timedelta(days=7)
                end = today
                mode_text = "за неделю"
            else:
                start = today - timedelta(days=30)
                end = today
                mode_text = "за месяц"

            basis_text = (
                "по дате начала (O)" if basis == "start" else "по дате окончания (P)"
            )

            header = (
                f"📋 Итоговые проверки {mode_text} {basis_text}\n"
                f"{start:%d.%m.%Y} — {end:%d.%m.%Y}"
            )
            text_out = build_final_checks_text_filtered(
                df,
                start_date=start,
                end_date=end,
                header=header,
                basis=basis,
            )
            await send_long_text(query.message.chat, text_out)
            await send_final_checks_xlsx_filtered(
                chat_id=query.message.chat.id,
                df=df,
                context=context,
                start_date=start,
                end_date=end,
                basis=basis,
            )
            context.user_data.pop("final_range_choice", None)
            return

        # пользовательский период
        if mode == "period":
            context.user_data["final_period"] = {
                "step": "start",
                "basis": basis,
            }
            context.user_data.pop("final_range_choice", None)
            await query.message.reply_text(
                "Введите дату начала периода (ДД.ММ.ГГГГ):"
            )
            return

        # на всякий случай
        context.user_data.pop("final_range_choice", None)
        await query.message.reply_text(
            "Что-то пошло не так. Попробуйте ещё раз выбрать режим."
        )
        return

    if data == "final_search_case":
        context.user_data["awaiting_final_case_search"] = True
        await query.message.reply_text(
            "Введите номер дела (формат 00-00-000000), который нужно найти "
            "в итоговых проверках:"
        )
        return


# -------------------------------------------------
# TEXT ROUTER
# -------------------------------------------------
async def text_router(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    chat = update.message.chat

    # Инспектор — пошаговый мастер
    if "inspector_form" in context.user_data:
        await inspector_process(update, context)
        return

    # Итоговые проверки — пользовательский период
    if context.user_data.get("final_period"):
        period = context.user_data["final_period"]
        step = period.get("step")
        basis = period.get("basis", "any")

        # ШАГ 1: ввод даты начала
        if step == "start":
            try:
                start_date = datetime.strptime(text, "%d.%m.%Y").date()
                if start_date.year < 2000 or start_date.year > 2100:
                    raise ValueError("year out of range")

                period["start_date"] = start_date
                period["step"] = "end"
                context.user_data["final_period"] = period
                await update.message.reply_text(
                    "Введите дату окончания периода (ДД.ММ.ГГГГ):"
                )
            except Exception:
                await update.message.reply_text(
                    "Дата начала в неверном формате.\n"
                    "Введите в виде ДД.ММ.ГГГГ (например, 05.01.2025)."
                )
            return

        # ШАГ 2: ввод даты окончания
        if step == "end":
            try:
                end_date = datetime.strptime(text, "%d.%m.%Y").date()
                if end_date.year < 2000 or end_date.year > 2100:
                    raise ValueError("year out of range")

                start_date = period.get("start_date")
                if start_date and end_date < start_date:
                    await update.message.reply_text(
                        "Дата окончания раньше даты начала.\n"
                        "Введите корректную дату окончания (ДД.ММ.ГГГГ)."
                    )
                    return

                df = get_final_checks_df()
                if df is None:
                    await update.message.reply_text(
                        "Не удалось открыть таблицу итоговых проверок."
                    )
                    context.user_data.pop("final_period", None)
                    return

                basis_text = (
                    "по дате начала (O)" if basis == "start" else "по дате окончания (P)"
                )
                header = (
                    f"📋 Итоговые проверки {basis_text} "
                    f"за период {start_date:%d.%m.%Y} — {end_date:%d.%m.%Y}"
                )
                text_out = build_final_checks_text_filtered(
                    df,
                    start_date=start_date,
                    end_date=end_date,
                    header=header,
                    basis=basis,
                )
                await send_long_text(chat, text_out)
                await send_final_checks_xlsx_filtered(
                    chat_id=chat.id,
                    df=df,
                    context=context,
                    start_date=start_date,
                    end_date=end_date,
                    basis=basis,
                )
                context.user_data.pop("final_period", None)
            except Exception:
                await update.message.reply_text(
                    "Дата окончания в неверном формате.\n"
                    "Введите в виде ДД.ММ.ГГГГ (например, 12.12.2025)."
                )
            return

    # Комментарий к доработке графика
    if context.user_data.get("awaiting_rework_comment"):
        info = context.user_data.pop("awaiting_rework_comment")
        version = info["version"]
        approver = info["approver"]
        comment = text
        update_schedule_approval_status(version, approver, "rework", comment)
        await update.message.reply_text(
            "Комментарий сохранён. График помечен как отправленный на доработку."
        )
        return

    # Ввод списка согласующих
    if context.user_data.get("awaiting_approvers_input"):
        info = context.user_data.pop("awaiting_approvers_input")
        version = info["version"]

        raw = text.replace(",", " ").split()
        approvers: List[str] = []
        for token in raw:
            token = token.strip()
            if not token:
                continue
            if not token.startswith("@"):
                token = "@" + token
            approvers.append(token)
        approvers = list(dict.fromkeys(approvers))

        if not approvers:
            await update.message.reply_text("Не найдено ни одного юзернейма.")
            return

        set_current_approvers_for_version(approvers, version)

        lines = [
            "График на новую неделю, необходимо согласовать.",
            f"Версия: {version}",
            "",
            "Согласующие:",
        ]
        for a in approvers:
            lines.append(f"• {a}")

        kb = InlineKeyboardMarkup(
            [
                [
                    InlineKeyboardButton(
                        f"✅ Согласовать ({a})", callback_data=f"schedule_approve:{a}"
                    ),
                    InlineKeyboardButton(
                        f"✏️ На доработку ({a})",
                        callback_data=f"schedule_rework:{a}",
                    ),
                ]
                for a in approvers
            ]
        )

        text_to_send = "\n".join(lines)

        await chat.send_message(text_to_send, reply_markup=kb)

        if SCHEDULE_NOTIFY_CHAT_ID is not None:
            try:
                await context.bot.send_message(
                    chat_id=SCHEDULE_NOTIFY_CHAT_ID,
                    text=text_to_send,
                    reply_markup=kb,
                )
            except Exception as e:
                log.error(
                    "Не удалось отправить уведомление в чат SCHEDULE_NOTIFY_CHAT_ID=%s: %s",
                    SCHEDULE_NOTIFY_CHAT_ID,
                    e,
                )

        await update.message.reply_text("Согласующие сохранены и уведомлены.")
        return

    # Поиск по номеру дела в замечаниях
    if context.user_data.get("awaiting_case_search"):
        context.user_data.pop("awaiting_case_search", None)
        case_no = text.strip()
        df = get_remarks_df_current()
        if df is None:
            await update.message.reply_text(
                "Не удалось открыть файл замечаний. Проверьте доступ к таблице."
            )
            return
        out_text = build_case_cards_text(df, case_no)
        await send_long_text(chat, out_text)
        return

    # Поиск по номеру дела в итоговых проверках
    if context.user_data.get("awaiting_final_case_search"):
        context.user_data.pop("awaiting_final_case_search", None)
        case_no = text.strip()
        df = get_final_checks_df()
        if df is None:
            await update.message.reply_text(
                "Не удалось открыть таблицу итоговых проверок."
            )
            return
        header = f"📋 Итоговые проверки по номеру дела: {case_no}"
        text_out = build_final_checks_text_filtered(
            df, case_no=case_no, header=header
        )
        await send_long_text(chat, text_out)
        await send_final_checks_xlsx_filtered(
            chat_id=chat.id, df=df, context=context, case_no=case_no
        )
        return

    low = text.lower()

    if low == "📅 график".lower():
        settings = get_schedule_state()
        is_adm = is_admin(update.effective_user.id)
        msg = build_schedule_text(is_adm, settings)
        user_username = update.effective_user.username or ""
        user_tag = f"@{user_username}" if user_username else None
        kb = build_schedule_inline(is_adm, settings, user_tag=user_tag)
        msg_full = (
            "📅 Раздел «График выездов»\n\n"
            "• Смотреть текущий статус согласования\n"
            "• Обновить данные из общей таблицы\n"
            "• Скачать красиво оформленный Excel-файл\n\n"
            "Если вы входите в список согласующих, ниже будут кнопки "
            "«Согласовать» и «На доработку».\n\n"
            f"{msg}"
        )
        await update.message.reply_text(msg_full, reply_markup=kb)
        return

    if low == "📝 замечания".lower():
        kb = remarks_menu_inline()
        msg = (
            "📝 Раздел «Замечания»\n\n"
            "Здесь доступны:\n"
            "• 🔎 поиск по номеру дела (столбец I);\n"
            "• 🏗 ОНзС — выбор 1–12, список дел (Номер дела (I) + адрес) и отдельный просмотр неустранённых;\n"
            "• 📥 открыть общий файл таблицы.\n\n"
            "Выберите нужное действие:"
        )
        await update.message.reply_text(msg, reply_markup=kb)
        return

    if low in ("инспектор", "👮 инспектор"):
        kb = inspector_menu_inline()
        msg = (
            "👮‍♂️ Раздел «Инспектор»\n\n"
            "Здесь можно:\n"
            "• ➕ добавить выезд инспектора;\n"
            "• 📋 посмотреть последние выезды;\n"
            "• 📥 скачать отдельный Excel с выездами;\n"
            "• 🔄 обнулить список выездов (кнопка «Обновить»).\n\n"
            "Выберите действие кнопками ниже."
        )
        await update.message.reply_text(msg, reply_markup=kb)
        return

    if low == "📈 аналитика".lower():
        conn = get_db()
        c = conn.cursor()
        c.execute(
            """SELECT version, approver, status, comment, decided_at, requested_at
               FROM schedule_approvals
               ORDER BY version DESC, approver"""
        )
        rows = c.fetchall()
        conn.close()

        if not rows:
            await update.message.reply_text("Пока нет данных по согласованию графика.")
            return

        by_ver: Dict[int, List[sqlite3.Row]] = {}
        for r in rows:
            by_ver.setdefault(r["version"], []).append(r)

        lines: List[str] = ["📈 Аналитика по согласованию графика:", ""]

        for ver in sorted(by_ver.keys(), reverse=True):
            approvals = by_ver[ver]
            header = build_schedule_header(ver, approvals)
            lines.append("")
            lines.append(header + ":")
            for r in approvals:
                appr = r["approver"]
                status = r["status"] or "pending"
                decided = _format_dt(r["decided_at"])
                requested = _format_dt(r["requested_at"])
                comment = r["comment"] or ""

                if status == "pending":
                    lines.append(f"• {appr} — ожидает, запрошено {requested}")
                elif status == "approved":
                    lines.append(f"• {appr} — Согласовано {decided} ✅")
                elif status == "rework":
                    if comment:
                        lines.append(
                            f"• {appr} — На доработку {decided} (Комментарий: {comment})"
                        )
                    else:
                        lines.append(f"• {appr} — На доработку {decided}")

        await send_long_text(chat, "\n".join(lines))
        return

    if low == "итоговые проверки":
        kb = final_checks_menu_inline()
        msg = (
            "📋 Раздел «Итоговые проверки»\n\n"
            "Вы можете:\n"
            "• посмотреть проверки за последнюю неделю;\n"
            "• за последний месяц;\n"
            "• указать свой период дат;\n"
            "• выполнить поиск по номеру дела.\n\n"
            "Выберите нужный вариант кнопками ниже."
        )
        await update.message.reply_text(msg, reply_markup=kb)
        return

    await update.message.reply_text(
        "Я вас не понял. Выберите пункт меню или нажмите /start.",
        reply_markup=main_menu(),
    )


# -------------------------------------------------
# DOCUMENT HANDLER
# -------------------------------------------------
async def document_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "Загрузка файлов через бота отключена. Используйте общую Google-таблицу."
    )


# -------------------------------------------------
# START / HELP
# -------------------------------------------------
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    msg = (
        "Добро пожаловать в бота отдела СОТ.\n\n"
        "Основные разделы:\n"
        "• 📅 График — согласование графика выездов\n"
        "• 📝 Замечания — поиск по номеру дела, ОНзС и статусы «нет»\n"
        "• Инспектор — выезды инспектора\n"
        "• Итоговые проверки — перечень итоговых проверок по отдельной таблице\n"
        "• 📈 Аналитика — история согласований\n\n"
        "Выберите раздел с помощью кнопок ниже."
    )
    await update.message.reply_text(msg, reply_markup=main_menu())


async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    msg = (
        "Справка по боту СОТ:\n\n"
        "📅 График — показать статус согласования, обновить, скачать Excel.\n"
        "📝 Замечания — поиск по номеру дела (I), работа с ОНзС и просмотр статусов «нет».\n"
        "Инспектор — добавление и выгрузка выездов инспектора.\n"
        "Итоговые проверки — список и выгрузка итоговых проверок за период или по делу.\n"
        "📈 Аналитика — история согласований по версиям графика.\n"
    )
    await update.message.reply_text(msg, reply_markup=main_menu())


# -------------------------------------------------
# MAIN
# -------------------------------------------------
def main():
    if not BOT_TOKEN:
        log.error("BOT_TOKEN не задан.")
        raise SystemExit("Укажите BOT_TOKEN в переменных окружения.")

    init_db()

    app = Application.builder().token(BOT_TOKEN).build()

    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("help", help_command))

    app.add_handler(CallbackQueryHandler(callback_handler))

    app.add_handler(MessageHandler(filters.Document.ALL, document_handler))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, text_router))

    log.info("Бот запущен...")
    app.run_polling()


if __name__ == "__main__":
    main()
